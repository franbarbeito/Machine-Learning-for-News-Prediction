from main_ui import *
from os import listdir, mkdir, linesep
from os.path import expanduser, isfile, join, split
from PyQt5.QtWidgets import QFileDialog, QTableWidgetItem
from nltk.corpus import stopwords
from nltk.stem import SnowballStemmer
from nltk import word_tokenize
from collections import Counter
from sklearn.tree import DecisionTreeClassifier # Import Decision Tree Classifier
from sklearn.neighbors import KNeighborsClassifier
from sklearn.model_selection import train_test_split # Import train_test_split function
from sklearn import metrics #Import scikit-learn metrics module for accuracy calculation
from joblib import dump, load
from sklearn.naive_bayes import MultinomialNB
from sklearn.metrics import confusion_matrix
from sklearn.feature_extraction.text import TfidfTransformer
from pickle import dump, load
from pandas import DataFrame
from glob import glob
import ctypes
from random import randrange
from openpyxl import Workbook
from shutil import copy
from tkinter import messagebox


class Modelo():

    def __init__(self, clasificador, precision, acierto, matrizResultados, algoritmo):
        self.clasificador = clasificador
        self.precision = precision
        self.acierto = acierto
        self.matrizResultados = matrizResultados
        self.algoritmo = algoritmo


class MainWindow(QtWidgets.QMainWindow, Ui_MainWindow):

    dirModelo = None
    dirOdio = None
    dirNoOdio = None
    dirRutaNoticias = None
    noticiasTesting = dict()
    listaPalabrasContadasOdio = []
    listaPalabrasContadasNoOdio = []
    listaPalabrasContadasSinMarcar = []
    listaPalabras = []
    listaPalabrasDiccionario = []
    modelo = None
    listaPalabrasCompleta = []
    listaNoticias = []
    modeloGenerado = None

    

    #Inicializamos el stopwords
    stop_words = set(stopwords.words('spanish'))

    #Inicializamos el stemmer
    stemmer = SnowballStemmer('spanish')



    def __init__(self, *args, **kwargs):
        #Inicializacion de la ventana y listeners
        QtWidgets.QMainWindow.__init__(self, *args, **kwargs)
        self.setupUi(self)
        #Signals
        self.btnRutaOdio.clicked.connect(lambda: self.abrirRuta(0))
        self.btnRutaNoOdio.clicked.connect(lambda: self.abrirRuta(1))
        self.btnRutaNoticias.clicked.connect(lambda: self.abrirRuta(2))
        self.btnGuardarModelo.clicked.connect(lambda: self.guardarModelo(self.modeloGenerado))
        self.btnRutaModelo.clicked.connect(self.abrirModelo)
        self.btnGenerarModelo.clicked.connect(self.generarModelo)
        self.btnExportar.clicked.connect(lambda: self.exportarResultado(self.noticiasTesting))
        self.btnClasificar.clicked.connect(self.clasificarNoticias)

        #Inicializacion diccionario castellano stemmed
        self.listaPalabrasDiccionario = self.generarListaTXT()
        
    
    def abrirRuta(self, i):
        #Guardamos en la variable la ruta seleccionada a traves de la ventana.
        my_dir = str(QFileDialog.getExistingDirectory(self, "Abre una carpeta", expanduser("~"), QFileDialog.ShowDirsOnly))

        #Dependiendo del boton presionado se guardara la ruta en diferentes variables
        if i == 0:
            if len(my_dir) > 50:
                self.lblRutaOdio.setText("..." + my_dir[-51:])
            else:
                self.lblRutaOdio.setText(my_dir)

            self.dirOdio = my_dir

        elif i == 1: 
            if len(my_dir) > 50:
                self.lblRutaNoOdio.setText("..." + my_dir[-51:])
            else:
                self.lblRutaNoOdio.setText(my_dir)

            self.dirNoOdio = my_dir

        elif i == 2: 
            if len(my_dir) > 50:
                self.lblRutaNoticias.setText("..." + my_dir[-51:])
            else:
                self.lblRutaNoticias.setText(my_dir)

            self.dirRutaNoticias = my_dir

    def abrirModelo(self):
        #Obtenemos la ruta del modelo
        modelo = QFileDialog.getOpenFileName(self, 'Abrir modelo', expanduser("~"), "Modelo (*.model)")
        if modelo[0] != "":
            #Si hemos obtenido una ruta, cargamos el modelo en la variable y mostramos las estadisticas
            self.lblRutaModelo.setText(modelo[0])
            self.dirModelo = modelo[0]
            self.modelo = self.cargarModelo(self.dirModelo)
            self.lblAlgoritmoTesting.setText(self.modelo.algoritmo)
            self.lblAciertoTesting.setText(str(self.modelo.acierto) + "%")
            self.lblPrecisionTesting.setText(str(self.modelo.precision) + "%")
            self.mostrarResultadosTabla(self.modelo.matrizResultados, self.tableEstadisticasTesting)
    

    def getFicherosDirectorio(self, dir):
        #Crea una lista con la ruta de todos los ficheros
        ficheros = glob(dir + "/*.txt")
        return ficheros


    def generarModelo(self):
        self.progressBar.setProperty("value", 0)
        #Comprobacion de que todas las rutas se han seleccionado
        if self.dirOdio == None or self.dirNoOdio == None:
            #Mensaje de error
            ctypes.windll.user32.MessageBoxW(0, "Debes elegir una ruta para todos las noticias", "Error al recuperar noticias", 0)
        else:
            #Obtenemos las rutas de los archivos
            self.progressBar.setProperty("value", 5)
            noticiasNoOdio = self.getFicherosDirectorio(self.dirNoOdio)
            noticiasOdio = self.getFicherosDirectorio(self.dirOdio)
            print(len(noticiasOdio))


            #Procesamiento de texto
            listaPalabrasContadasOdio = self.getPalabrasContadas(noticiasOdio, 1)
            listaPalabrasContadasNoOdio = self.getPalabrasContadas(noticiasNoOdio, 0)
            listaPalabrasContadas = listaPalabrasContadasOdio + listaPalabrasContadasNoOdio
            self.progressBar.setProperty("value", 10)

            #Creamos el dataframe que le pasaremos al modelo con todas las palabras en las columnas y una fila por noticia 
            df = DataFrame(listaPalabrasContadas, columns=self.listaPalabrasDiccionario).fillna(0)
            

            #En X tenemos todas las palabras y su cuenta, es decir los datos que vamos a necesitar para predecir
            X = df.drop("esOdio", axis=1)

            #En y tenemos la columna que queremos que prediga
            y = df["esOdio"]

            self.progressBar.setProperty("value", 20)
            #Divide los datos en training y testing
            X_train, X_test, y_train, y_test = train_test_split(X, y, test_size=0.2, random_state=1) # 80% training y 20% test

            #Creamos el clasificador dependiendo del algoritmo elegido y lo entrenamos
            #Arbol de decision
            if self.cbAlgoritmo.currentIndex() == 0:
                clf = DecisionTreeClassifier()    
                clf = clf.fit(X_train,y_train)
            
            #K-nn Neighbours
            elif self.cbAlgoritmo.currentIndex() == 1:
                clf = KNeighborsClassifier(n_neighbors=5)   
                clf = clf.fit(X_train,y_train)

            #Naive Bayes
            elif self.cbAlgoritmo.currentIndex() == 2:
                clf = MultinomialNB()   
                clf = clf.fit(X_train,y_train)

            #Guardamos en variables los resultados de la prediccion
            y_pred = clf.predict(X_test)
            algoritmo = self.cbAlgoritmo.currentText()
            acierto = round(metrics.accuracy_score(y_test, y_pred)*100, 2)
            precision = round(metrics.precision_score(y_test, y_pred)*100, 2)
            self.progressBar.setProperty("value", 50)

            #Imprimimos el algoritmo utilizado en la etiqueta
            self.lblAlgoritmo.setText(algoritmo)

            #Imprimimos el % de acierto en la etiqueta
            self.lblAcierto.setText(str(acierto) + "%")

            #Imprimimos el % de precision en la etiqueta
            self.lblPrecision.setText(str(precision) + "%")
            
            self.progressBar.setProperty("value", 75)
            #Obtenemos la matriz de resultados e imprimimos resultados en la tabla
            matrizResultados = confusion_matrix(y_test, y_pred)

            self.mostrarResultadosTabla(matrizResultados, self.tableEstadisticas)
            
            #Guardamos el modelo generado
            self.modeloGenerado = Modelo(clf, precision, acierto, matrizResultados, algoritmo)
            self.progressBar.setProperty("value", 100)

    

    def mostrarResultadosTabla(self, matrizResultados, tableEstadisticas):
        tableEstadisticas.setItem(0, 0, QTableWidgetItem(str(matrizResultados[0][0])))
        tableEstadisticas.setItem(0, 1, QTableWidgetItem(str(matrizResultados[0][1])))
        tableEstadisticas.setItem(1, 0, QTableWidgetItem(str(matrizResultados[1][0])))
        tableEstadisticas.setItem(1, 1, QTableWidgetItem(str(matrizResultados[1][1])))
        tableEstadisticas.setItem(2, 0, QTableWidgetItem(str(round(matrizResultados[0][0]/(matrizResultados[0][0]+matrizResultados[1][0]), 2)*100) + "%"))
        tableEstadisticas.setItem(2, 1, QTableWidgetItem(str(round(matrizResultados[1][1]/(matrizResultados[0][1]+matrizResultados[1][1]), 2)*100) + "%"))
        tableEstadisticas.setItem(0, 2, QTableWidgetItem(str(round(matrizResultados[0][0]/(matrizResultados[0][0]+matrizResultados[0][1]), 2)*100) + "%"))
        tableEstadisticas.setItem(1, 2, QTableWidgetItem(str(round(matrizResultados[1][1]/(matrizResultados[1][0]+matrizResultados[1][1]), 2)*100) + "%"))


    #Funcion para clasificar las noticias con el modelo cargado y mostrarlas en la lista
    def clasificarNoticias(self):

        #Limpiamos el diccionario para no mezclar resultados
        self.noticiasTesting.clear()

        #Si tenemos un modelo y ruta seleccionados comenzamos el procesamiento de las noticias
        if self.modelo and self.dirRutaNoticias:
            noticias = self.getFicherosDirectorio(self.dirRutaNoticias)
            listaPalabrasContadas = self.getPalabrasContadas(noticias)

            counterTotal = Counter()
            for counter in listaPalabrasContadas:
                counterTotal += counter

            listaPalabras = []
            for palabra in counterTotal:
                listaPalabras.append(palabra)
            
            listaPalabras += self.listaPalabrasCompleta

            listaPalabrasFinal = []

            for palabra in listaPalabras:
                if palabra not in listaPalabrasFinal:
                    listaPalabrasFinal.append(palabra)
            
            dfNoticias = DataFrame(listaPalabrasContadas, columns=self.listaPalabrasDiccionario).fillna(0)
            print(dfNoticias)

            X = dfNoticias.drop("esOdio", axis=1)

            y_pred = self.modelo.clasificador.predict(X)
            


            #Creamos un diccionario con una clave por noticia y su valor sera la prediccion
            for noticia in noticias:
                prediccion = "No odio"

                if y_pred[noticias.index(noticia)] == 1:
                    prediccion = "Odio"

                print("Guardando: " + noticia + " --> " + prediccion)
                self.noticiasTesting[noticia] = prediccion
            
            self.mostrarResultadoNoticias(self.noticiasTesting)
        
        else:
            if self.modelo == None:
                ctypes.windll.user32.MessageBoxW(0, "Debes elegir un modelo", "Error al recuperar modelo", 0)
            
            if self.dirRutaNoticias == None:
                ctypes.windll.user32.MessageBoxW(0, "Debes elegir una ruta para las noticias", "Error al recuperar noticias", 0)
            




    def getPalabrasNoticia(self, noticia, utf=None):

        #Abrimos la noticia y guardamos en raw el texto plano
        if utf:
            f = open(noticia, encoding='utf-8')
        else:
            f = open(noticia)
        raw = f.read()

        #Tokenizamos el texto guardandolo en una lista
        tokens = word_tokenize(raw)

        #Quitamos los elementos que no sean palabras
        filteredAlNum = [w.lower() for w in tokens if w.isalnum()]

        #Quitamos los elementos que sean preposiciones, determinantes, etc (palabras que no aportan la informacion que necesitamos)
        filteredStopwords = [w for w in filteredAlNum if not w in self.stop_words]

        #Eliminamos los sufijos y prefijos de las palabras de la ultima lista
        filteredStem = [self.stemmer.stem(w) for w in filteredStopwords]

        return filteredStem
    
    def getPalabrasContadas(self, listaNoticias, esOdio=None):

        listaPalabrasContadas = []
        #Abrimos cada noticia y guardamos en un array las palabras de cada una
        for noticia in listaNoticias:
            listaPalabrasNoticia = self.getPalabrasNoticia(noticia)
            #Contamos todas las palabras que se repiten guardando las palabras y las veces que aparecen
            listaPalabrasContadasNoticia = Counter(listaPalabrasNoticia)
            if esOdio:
                #Añadimos una nueva pareja, que nos servira para tener este dato en el dataframe
                listaPalabrasContadasNoticia["esOdio"] = esOdio
            listaPalabrasContadas.append(listaPalabrasContadasNoticia)

        return listaPalabrasContadas

    #Funcion para exportar el resultado a un excel y ademas crear dos carpetas de clasificacion metiendo en cada una las noticias correspondientes
    def exportarResultado(self, dicNoticias):
        if dicNoticias:
            wb = Workbook()
            ruta = QFileDialog.getExistingDirectory(self, "Selecciona una carpeta", expanduser("~"), QFileDialog.ShowDirsOnly)
            if ruta != "":
                mkdir(ruta + "/Odio")
                mkdir(ruta + "/No odio")

                hoja = wb.active
                hoja.title = "Noticia-Valor"

                fila = 1 #Fila donde empezamos
                colNoticia = 1 #Columna donde guardamos los nombres de las noticias
                colOdio = 2 #Columna donde guardamos la prediccion

                for noticia in dicNoticias:
                    hoja.cell(column=colNoticia, row=fila, value=split(noticia)[1])
                    hoja.cell(column=colOdio, row=fila, value=dicNoticias[noticia])
                    if dicNoticias[noticia] == "Odio":
                        #Mover noticia a carpeta ruta/Odio
                        print("Moviendo a odio: " + noticia)
                        copy(noticia, ruta + "/Odio/" + split(noticia)[1])
                    else:
                        #Mover noticia a carpeta ruta/No_odio
                        print("Moviendo a no odio: " + noticia)
                        copy(noticia, ruta + "/No odio/" + split(noticia)[1])
                    fila+=1

                wb.save(filename=ruta + "/resultado.xlsx")
                messagebox.showinfo('Exportado', 'Los resultados se han exportando con exito')
              
        
        else:
            ctypes.windll.user32.MessageBoxW(0, "Debes clasificar las noticias con un modelo primero", "Error al exportar resultado", 0)


    
    def guardarModelo(self, modelo):
        if modelo:
            filename = QFileDialog.getSaveFileName(self, caption="Guardar modelo", filter="Modelo (*.model)")
            if filename[0] != "":
                with open(filename[0], 'wb') as file:
                    dump(modelo, file)
        else:
            ctypes.windll.user32.MessageBoxW(0, "Debes generar un modelo", "Error al guardar modelo", 0)

    #Funcion para guardar en una variable el modelo cargado
    def cargarModelo(self, filename):
        with open(filename, 'rb') as file:
            pickle_model = load(file)

        return pickle_model

    def mostrarResultadoNoticias(self, dicPrediccion):
        #Establecemos el numero de filas
        self.tableResultados.setRowCount(len(dicPrediccion))
        i=0
        #Por cada noticia se muestra en la tabla la noticia y la prediccion
        for noticia in dicPrediccion:
            self.tableResultados.setItem(i, 0, QTableWidgetItem(split(noticia)[1]))
            self.tableResultados.setItem(i, 1, QTableWidgetItem(dicPrediccion[noticia]))
            i+=1
            
    
   

    #Funcion para recoger el diccionario lematizado y guardarlo en una lista
    def generarListaTXT(self):
        lista = []
        file = open("dicStemmed.txt", "r")
        for linea in file:
            lista.append(linea[:-1])
        
        file.close()
        return lista


   
    


    
if __name__ == "__main__":
    app = QtWidgets.QApplication([])
    window = MainWindow()
    window.show()
    app.exec_()